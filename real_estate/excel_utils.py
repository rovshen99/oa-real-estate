from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


def create_real_estate_excel(real_estates):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Home Ojar Azia"

    headers = ["№", "Ф.И.О.", "Тип", "Фактическая стоимость за 1м2 $", "Корпус", "№ подъезда", "Этаж", "Число комнат", "Общ. Площ. м2", "Фактическая стоимость $", "Расчет", "Дата оплаты", "Остаток", "Примечание"]
    header_fill = PatternFill(start_color="F4A460", end_color="F4A460", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.font = Font(name="Times New Roman", bold=True, size=10)
        cell.border = thin_border

    sheet.row_dimensions[1].height = 30

    row_num = 2
    for idx, real_estate in enumerate(real_estates, start=1):
        transactions = real_estate.transactions.all()
        if transactions:
            for txn_idx, transaction in enumerate(transactions):
                payments = transaction.payments.all()
                max_payments = max(1, len(payments))

                if max_payments > 1:
                    sheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num + max_payments - 1, end_column=1)
                    sheet.merge_cells(start_row=row_num, start_column=2, end_row=row_num + max_payments - 1, end_column=2)
                    sheet.merge_cells(start_row=row_num, start_column=3, end_row=row_num + max_payments - 1, end_column=3)
                    sheet.merge_cells(start_row=row_num, start_column=4, end_row=row_num + max_payments - 1, end_column=4)
                    sheet.merge_cells(start_row=row_num, start_column=5, end_row=row_num + max_payments - 1, end_column=5)
                    sheet.merge_cells(start_row=row_num, start_column=6, end_row=row_num + max_payments - 1, end_column=6)
                    sheet.merge_cells(start_row=row_num, start_column=7, end_row=row_num + max_payments - 1, end_column=7)
                    sheet.merge_cells(start_row=row_num, start_column=8, end_row=row_num + max_payments - 1, end_column=8)
                    sheet.merge_cells(start_row=row_num, start_column=9, end_row=row_num + max_payments - 1, end_column=9)
                    sheet.merge_cells(start_row=row_num, start_column=13, end_row=row_num + max_payments - 1, end_column=13)
                    sheet.merge_cells(start_row=row_num, start_column=14, end_row=row_num + max_payments - 1, end_column=14)

                for payment_idx, payment in enumerate(payments):
                    if payment_idx == 0:
                        sheet.cell(row=row_num, column=1, value=idx).alignment = Alignment(horizontal="center", vertical="center")
                        sheet.cell(row=row_num, column=2, value=transaction.buyer.name if transaction.buyer else "Оджар Азия").alignment = Alignment(horizontal="left", vertical="center")
                        sheet.cell(row=row_num, column=3, value=real_estate.type.name if real_estate.type else "").alignment = Alignment(horizontal="center", vertical="center")
                        sheet.cell(row=row_num, column=4, value=f"{real_estate.cost_per_sqm} {real_estate.currency}").alignment = Alignment(horizontal="center", vertical="center")
                        sheet.cell(row=row_num, column=5, value=real_estate.building).alignment = Alignment(horizontal="center", vertical="center")
                        sheet.cell(row=row_num, column=6, value=real_estate.entrance_number).alignment = Alignment(horizontal="center", vertical="center")
                        sheet.cell(row=row_num, column=7, value=real_estate.floor).alignment = Alignment(horizontal="center", vertical="center")
                        sheet.cell(row=row_num, column=8, value=real_estate.rooms_count).alignment = Alignment(horizontal="center", vertical="center")
                        sheet.cell(row=row_num, column=9, value=real_estate.total_area).alignment = Alignment(horizontal="center", vertical="center")
                        sheet.cell(row=row_num, column=10, value=f"{real_estate.total_cost} {real_estate.currency}").alignment = Alignment(horizontal="center", vertical="center")
                        sheet.cell(row=row_num, column=13, value=transaction.remaining_payment()).alignment = Alignment(horizontal="center", vertical="center")
                        sheet.cell(row=row_num, column=14, value=real_estate.description).alignment = Alignment(horizontal="center", vertical="center")

                    sheet.cell(row=row_num, column=11, value=f"{payment.amount} {real_estate.currency}" if payment else None).alignment = Alignment(horizontal="center", vertical="center")
                    sheet.cell(row=row_num, column=12, value=payment.date.strftime('%d/%m/%Y') if payment else None).alignment = Alignment(horizontal="center", vertical="center")
                    row_num += 1

        else:
            sheet.cell(row=row_num, column=1, value=idx).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row=row_num, column=2, value="Оджар Азия").alignment = Alignment(horizontal="left", vertical="center")
            sheet.cell(row=row_num, column=3, value=real_estate.type.name if real_estate.type else "").alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row=row_num, column=4, value=f"{real_estate.cost_per_sqm} {real_estate.currency}").alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row=row_num, column=5, value=real_estate.building).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row=row_num, column=6, value=real_estate.entrance_number).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row=row_num, column=7, value=real_estate.floor).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row=row_num, column=8, value=real_estate.rooms_count).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row=row_num, column=9, value=real_estate.total_area).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row=row_num, column=10, value=f"{real_estate.total_cost} {real_estate.currency}").alignment = Alignment(horizontal="center", vertical="center")
            sheet.merge_cells(start_row=row_num, start_column=11, end_row=row_num, end_column=13)
            sheet.cell(row=row_num, column=11).value = "СВОБОДНА"
            sheet.cell(row=row_num, column=11).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row=row_num, column=11).border = thin_border
            sheet.cell(row=row_num, column=14, value="").alignment = Alignment(horizontal="center", vertical="center")
            row_num += 1

    column_widths = [5, 20, 20, 20, 10, 10, 10, 10, 15, 20, 20, 20, 15, 20]
    for col_num, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(col_num)].width = width

    for row in sheet.iter_rows(min_row=2, max_row=row_num + 3, min_col=1, max_col=14):
        for cell in row:
            cell.border = thin_border
            cell.font = Font(name="Times New Roman", size=10)

    # Add summary at the bottom
    total_units = len(real_estates)
    sold_units = sum(1 for re in real_estates if re.transactions.exists())
    remaining_units = total_units - sold_units

    sheet.append(["", "", "", "", "", "", "", "", "", "", "", "", ""])
    sheet.append(["", "", f"ИТОГО: {total_units}"])
    sheet.append(["", "", f"ПРОДАНО: {sold_units}"])
    sheet.append(["", "", f"ОСТАТОК: {remaining_units}"])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def get_filename():
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"home_ojar_azia_{now}.xlsx"
